"""Module with the functions to work with the worksheets."""

import re
from typing import Any, Union

from openpyxl import load_workbook

import standards.config as cfg


def is_standard_compared(standard: Any) -> bool:
    """Check if the standard has already been compared.
    A compared standard has the cells 6, 7 and 8 filled (columns F, G and H).
    """
    return standard[4] is not None and standard[5] is not None and standard[6] is not None


def get_original_standards(path: str, worksheet: str) -> list[dict[str, Union[str, None]]]:
    """Gets the list with the original standards."""
    original_workbook = load_workbook(path)
    original_worksheet = original_workbook[cfg.original_standards_ws[cfg.get_worksheet_index(
        worksheet)]]
    original_rows = original_worksheet.iter_rows(
        min_row=4, min_col=2, max_col=8, values_only=True)

    original_standards: list[dict[str, Union[str, None]]] = []
    for row in original_rows:
        if row[0]:
            matches = re.match(
                r"^[A-Za-z]+[0-9]*(\.[0-9]+)*[a-z]*\.?[a-z]*$", row[0].strip())
            print("Row: ", row[0], "Match: ", matches)

        if row[0] and row[1] and row[0] != "No." and matches:
            standard = {
                "id": row[0],
                "text": row[1],
                "level": row[2] if row[2] else None,
                "completed": is_standard_compared(row)
            }
            original_standards.append(standard)

    return original_standards


def get_new_standards(path: str) -> list[dict[str, Union[str, None]]]:
    """Gets the list with the new standards."""
    new_workbook = load_workbook(path)
    new_worksheet = new_workbook["Unified Standard"]
    new_rows = new_worksheet.iter_rows(
        min_row=4, min_col=1, max_col=3, values_only=True)

    new_standards: list[dict[str, Union[str, None]]] = []
    for row in new_rows:
        if row[0] and row[1] and row[0] != "Criterion #":
            standard = {
                "id": row[0],
                "text": row[1],
                "level": row[2] if row[2] else None
            }
            new_standards.append(standard)

    return new_standards


def update_standards(id_value: str, new_id: str, new_text: str, new_level: str,
                     worksheet_name: str, workbook_path: str
                     ) -> bool:
    """Update the standards with new values."""
    workbook = load_workbook(workbook_path)
    worksheet = workbook[worksheet_name]

    row_number = None

    for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=2):
        if row[0].value == id_value:
            row_number = row[0].row
            break

    if not row_number:
        return False

    # Update the values in the row.
    worksheet[f"F{row_number}"] = new_id
    worksheet[f"G{row_number}"] = new_text
    worksheet[f"H{row_number}"] = new_level

    try:
        workbook.save(workbook_path)
    except PermissionError:
        print("Please close the file first before updating the standards.")
        return False
    return True


def get_cell_number_from_value(value, workbook_name):
    workbook = load_workbook(workbook_name)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=4, min_col=1, max_col=1):
        if row[0].value == value:
            return row[0].row, row[0].coordinate.replace("A", "B")

    return None
